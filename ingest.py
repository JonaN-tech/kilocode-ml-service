from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook
from pypdf import PdfReader


@dataclass
class CommentRecord:
    comment_text: str
    post_url: Optional[str] = None
    comment_url: Optional[str] = None
    posted_from: Optional[str] = None
    feedback: Optional[str] = None


def _clean_text(s: str) -> str:
    s = s.replace("\r", "\n")
    s = re.sub(r"\n{3,}", "\n\n", s)
    s = re.sub(r"[ \t]{2,}", " ", s).strip()
    return s


def load_comments_from_xlsx(xlsx_path: Path) -> List[CommentRecord]:
    wb = load_workbook(filename=xlsx_path, read_only=True)
    sheet = wb.active

    # Read header row
    headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
    header_index = {h: i for i, h in enumerate(headers)}

    def safe_cell(row, name: str) -> str:
        idx = header_index.get(name)
        if idx is None:
            return ""
        value = row[idx]
        return "" if value is None else str(value).strip()

    out: List[CommentRecord] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        mention = _clean_text(safe_cell(row, "Mention text"))
        if not mention:
            continue

        post_url = safe_cell(row, "Reddit Thread") or None
        comment_url = safe_cell(row, "Link to the mention") or None
        posted_from = safe_cell(row, "Posted from account") or None
        feedback = safe_cell(row, "Feedback from Darko") or None

        if post_url and not post_url.startswith("http"):
            post_url = None
        if comment_url and not comment_url.startswith("http"):
            comment_url = None

        out.append(
            CommentRecord(
                comment_text=mention,
                post_url=post_url,
                comment_url=comment_url,
                posted_from=posted_from,
                feedback=feedback,
            )
        )

    # Deduplicate by normalized comment text
    seen = set()
    deduped: List[CommentRecord] = []

    for r in out:
        key = re.sub(r"\s+", " ", r.comment_text).strip().lower()
        if key in seen:
            continue
        seen.add(key)
        deduped.append(r)

    return deduped


def load_pdf_text(pdf_path: Path) -> str:
    reader = PdfReader(str(pdf_path))
    pages = []
    for p in reader.pages:
        pages.append(p.extract_text() or "")
    return _clean_text("\n\n".join(pages))
